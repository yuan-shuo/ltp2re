from openpyxl import load_workbook

# 加载Excel文件
workbook = load_workbook('data.xlsx')
sheet = workbook.active

# 读取Excel文件中的数据
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    item = {'subject': row[0], 'relation': row[1], 'object': row[2]}
    data.append(item)

# 打印读取的数据
print(data)
